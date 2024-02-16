from openpyxl import Workbook

# ESCRIBIR
""" workbook = wb """
wb = Workbook()

""" work sheet = ws """
ws = wb.active
ws["A1"] = 1000

wb.save(filename="curso.xlsx")

# LEER

from openpyxl import load_workbook

wb = load_workbook(filename= "curso.xlsx")
ws = wb["Sheet"]
print(ws["A1"].value)
